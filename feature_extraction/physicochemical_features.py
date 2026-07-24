from pathlib import Path
from typing import Mapping, Sequence

import numpy as np
from openpyxl import load_workbook


# All possible DNA dinucleotides.
EXPECTED_DINUCLEOTIDES = {
    f"{first}{second}"
    for first in "ACGT"
    for second in "ACGT"
}


def load_physicochemical_properties_di(
    file_path: str | Path,
) -> dict[str, np.ndarray]:
    """
    Load the normalized dinucleotide physicochemical property table.

    The expected Excel structure is:

        Column A: Dinucleotide
        Columns B-M: 12 physicochemical properties

    Returns:
        Dictionary mapping each dinucleotide to a float32 vector
        with 12 physicochemical properties.
    """

    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(
            f"Physicochemical property file not found: {file_path}"
        )

    workbook = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=True,
    )

    worksheet = workbook.active

    property_table: dict[str, np.ndarray] = {}

    try:
        # The first row contains the column headers.
        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):
            dinucleotide = str(row[0]).strip().upper()

            properties = np.asarray(
                row[1:13],
                dtype=np.float32,
            )

            if dinucleotide not in EXPECTED_DINUCLEOTIDES:
                raise ValueError(
                    f"Invalid dinucleotide '{dinucleotide}' "
                    f"at Excel row {row_number}."
                )

            if properties.shape != (12,):
                raise ValueError(
                    f"Expected 12 properties for {dinucleotide}, "
                    f"received {properties.shape[0]}."
                )

            if not np.isfinite(properties).all():
                raise ValueError(
                    f"Non-finite physicochemical value found "
                    f"for {dinucleotide}."
                )

            property_table[dinucleotide] = properties

    finally:
        workbook.close()

    missing_dinucleotides = (
        EXPECTED_DINUCLEOTIDES.difference(
            property_table.keys()
        )
    )

    if missing_dinucleotides:
        raise ValueError(
            "Missing dinucleotide properties: "
            + ", ".join(
                sorted(missing_dinucleotides)
            )
        )

    return property_table


def convertSampleToPhyChemVector_Di(
    sampleSeq: Sequence[str],
    PhyChemPropTable_Di: Mapping[
        str,
        Sequence[float],
    ],
    expected_length: int = 501,
    unknown_strategy: str = "zero",
) -> np.ndarray:
    """
    Convert DNA sequences into dinucleotide physicochemical matrices.

    This function follows the original Yeast Promoter implementation:
    every adjacent dinucleotide is represented using 12
    physicochemical properties.

    Args:
        sampleSeq:
            DNA sequences. Each sequence must contain 501 bases.

        PhyChemPropTable_Di:
            Dictionary containing 12 values for every dinucleotide.

        expected_length:
            Expected DNA sequence length. The DeepMeth input length
            is currently fixed to 501 bp.

        unknown_strategy:
            Handling strategy for dinucleotides containing N.

            "zero":
                Use a zero vector for unknown dinucleotides.

            "error":
                Raise an error when an unknown dinucleotide is found.

    Returns:
        NumPy array with shape:

            [number_of_samples, 12, expected_length - 1]

        For 501 bp sequences:

            [number_of_samples, 12, 500]
    """

    if len(sampleSeq) == 0:
        raise ValueError(
            "sampleSeq must contain at least one DNA sequence."
        )

    sequences = [
        str(sequence).strip().upper()
        for sequence in sampleSeq
    ]

    if unknown_strategy not in {"zero", "error"}:
        raise ValueError(
            "unknown_strategy must be either 'zero' or 'error'."
        )

    for sample_index, sequence in enumerate(sequences):
        if len(sequence) != expected_length:
            raise ValueError(
                f"Sequence {sample_index} has length "
                f"{len(sequence)}; expected {expected_length}."
            )

        invalid_bases = set(sequence).difference(
            {"A", "C", "G", "T", "N"}
        )

        if invalid_bases:
            raise ValueError(
                f"Sequence {sample_index} contains invalid bases: "
                + ", ".join(sorted(invalid_bases))
            )

    physicochemical_matrix = np.zeros(
        (
            len(sequences),
            12,
            expected_length - 1,
        ),
        dtype=np.float32,
    )

    for sample_number, sequence in enumerate(sequences):
        for position in range(expected_length - 1):
            dinucleotide = sequence[
                position : position + 2
            ]

            if dinucleotide in PhyChemPropTable_Di:
                physicochemical_matrix[
                    sample_number,
                    :,
                    position,
                ] = np.asarray(
                    PhyChemPropTable_Di[dinucleotide],
                    dtype=np.float32,
                )

            elif "N" in dinucleotide:
                if unknown_strategy == "zero":
                    # The matrix was initialized with zeros,
                    # so no assignment is required.
                    continue

                raise ValueError(
                    f"Unknown dinucleotide '{dinucleotide}' "
                    f"found in sequence {sample_number} "
                    f"at position {position}."
                )

            else:
                raise ValueError(
                    f"Dinucleotide '{dinucleotide}' does not exist "
                    "in the physicochemical property table."
                )

    return physicochemical_matrix



    