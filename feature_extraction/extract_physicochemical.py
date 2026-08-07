from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Mapping, Sequence

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import numpy as np
import pyarrow.parquet as pq
from openpyxl import load_workbook

from config.project_config import (
    DATASET_DIR,
    PHYSICOCHEMICAL_FEATURES_DIR,
    PHYSICOCHEMICAL_PROPERTY_FILE,
    PHYSICOCHEMICAL_SHARD_SIZE,
    SEQUENCE_LENGTH,
)

SPLIT_NAMES = ("train", "validation", "test")
COLUMNS_TO_READ = ["chrom", "canonical_position", "sequence", "label"]

# All possible DNA dinucleotides.
EXPECTED_DINUCLEOTIDES = {f"{first}{second}" for first in "ACGT" for second in "ACGT"}


def load_physicochemical_properties_di(file_path: str | Path) -> dict[str, np.ndarray]:
  
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"Physicochemical property file not found: {file_path}")

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    worksheet = workbook.active

    property_table: dict[str, np.ndarray] = {}

    try:
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            dinucleotide = str(row[0]).strip().upper()
            properties = np.asarray(row[1:13], dtype=np.float32)

            if dinucleotide not in EXPECTED_DINUCLEOTIDES:
                raise ValueError(f"Invalid dinucleotide '{dinucleotide}' at Excel row {row_number}.")

            if properties.shape != (12,):
                raise ValueError(f"Expected 12 properties for {dinucleotide}, received {properties.shape[0]}.")

            if not np.isfinite(properties).all():
                raise ValueError(f"Non-finite physicochemical value found for {dinucleotide}.")

            property_table[dinucleotide] = properties
    finally:
        workbook.close()

    missing_dinucleotides = EXPECTED_DINUCLEOTIDES.difference(property_table.keys())

    if missing_dinucleotides:
        raise ValueError("Missing dinucleotide properties: " + ", ".join(sorted(missing_dinucleotides)))

    return property_table


UNKNOWN_DINUCLEOTIDE_CODE = 255  # sentinel for any pair involving "N"


def dinucleotide_codes(property_table: Mapping[str, Sequence[float]]) -> list[str]:
    """The 16 dinucleotides in a fixed order - index i is code i everywhere below."""
    return sorted(property_table.keys())


def build_property_matrix_by_code(property_table: Mapping[str, Sequence[float]]) -> np.ndarray:
  
    codes = dinucleotide_codes(property_table)
    return np.stack([np.asarray(property_table[d], dtype=np.float32) for d in codes])


def _build_code_lookup(property_table: Mapping[str, Sequence[float]]) -> np.ndarray:
   
    lookup_table = np.full((128, 128), UNKNOWN_DINUCLEOTIDE_CODE, dtype=np.uint8)

    for code, dinucleotide in enumerate(dinucleotide_codes(property_table)):
        first_code, second_code = ord(dinucleotide[0]), ord(dinucleotide[1])
        lookup_table[first_code, second_code] = code

    return lookup_table


def _build_dinucleotide_lookup(
    property_table: Mapping[str, Sequence[float]],
) -> tuple[np.ndarray, np.ndarray]:
   
    lookup_table = np.zeros((128, 128, 12), dtype=np.float32)
    known_mask = np.zeros((128, 128), dtype=bool)

    for dinucleotide, properties in property_table.items():
        first_code, second_code = ord(dinucleotide[0]), ord(dinucleotide[1])
        lookup_table[first_code, second_code] = properties
        known_mask[first_code, second_code] = True

    return lookup_table, known_mask


def convertSampleToPhyChemVector_Di(
    sampleSeq: Sequence[str],
    PhyChemPropTable_Di: Mapping[str, Sequence[float]],
    expected_length: int = SEQUENCE_LENGTH,
    unknown_strategy: str = "zero",
) -> np.ndarray:
   
    if len(sampleSeq) == 0:
        raise ValueError("sampleSeq must contain at least one DNA sequence.")

    sequences = [str(sequence).strip().upper() for sequence in sampleSeq]

    if unknown_strategy not in {"zero", "error"}:
        raise ValueError("unknown_strategy must be either 'zero' or 'error'.")

    for sample_index, sequence in enumerate(sequences):
        if len(sequence) != expected_length:
            raise ValueError(f"Sequence {sample_index} has length {len(sequence)}; expected {expected_length}.")

        invalid_bases = set(sequence).difference({"A", "C", "G", "T", "N"})

        if invalid_bases:
            raise ValueError(
                f"Sequence {sample_index} contains invalid bases: " + ", ".join(sorted(invalid_bases))
            )

    joined_sequences = "".join(sequences).encode("ascii")
    char_codes = np.frombuffer(joined_sequences, dtype=np.uint8).reshape(len(sequences), expected_length)

    first_codes = char_codes[:, :-1]
    second_codes = char_codes[:, 1:]

    lookup_table, known_mask = _build_dinucleotide_lookup(PhyChemPropTable_Di)
    physicochemical_matrix = lookup_table[first_codes, second_codes]

    if unknown_strategy == "error":
        is_known = known_mask[first_codes, second_codes]
        contains_n = (first_codes == ord("N")) | (second_codes == ord("N"))
        invalid_positions = ~is_known & ~contains_n

        if invalid_positions.any():
            sample_index, position = (int(index[0]) for index in np.nonzero(invalid_positions))
            dinucleotide = sequences[sample_index][position : position + 2]
            raise ValueError(
                f"Dinucleotide '{dinucleotide}' does not exist in the physicochemical property table."
            )

    return physicochemical_matrix.transpose(0, 2, 1).astype(np.float32, copy=False)


def encode_sequences_to_codes(
    sequences: Sequence[str],
    property_table: Mapping[str, Sequence[float]],
    expected_length: int = SEQUENCE_LENGTH,
) -> np.ndarray:
   
    sequences = [str(sequence).strip().upper() for sequence in sequences]

    joined_sequences = "".join(sequences).encode("ascii")
    char_codes = np.frombuffer(joined_sequences, dtype=np.uint8).reshape(len(sequences), expected_length)

    code_lookup = _build_code_lookup(property_table)

    return code_lookup[char_codes[:, :-1], char_codes[:, 1:]]


def expand_codes_to_matrix(
    codes: np.ndarray,
    property_table: Mapping[str, Sequence[float]],
) -> np.ndarray:
    
    property_matrix_by_code = build_property_matrix_by_code(property_table)
    # Row UNKNOWN_DINUCLEOTIDE_CODE -> append a zero row so it maps cleanly.
    lookup_with_zero_row = np.vstack([property_matrix_by_code, np.zeros((1, 12), dtype=np.float32)])

    safe_codes = np.where(codes == UNKNOWN_DINUCLEOTIDE_CODE, len(property_matrix_by_code), codes)

    return lookup_with_zero_row[safe_codes].transpose(0, 2, 1)


def process_split(split_name: str, parquet_path: Path, property_table: dict) -> dict:
    output_dir = PHYSICOCHEMICAL_FEATURES_DIR / split_name
    output_dir.mkdir(parents=True, exist_ok=True)

    parquet_file = pq.ParquetFile(parquet_path)

    shard_records = []
    total_rows = 0
    shard_index = 0

    for record_batch in parquet_file.iter_batches(
        batch_size=PHYSICOCHEMICAL_SHARD_SIZE,
        columns=COLUMNS_TO_READ,
    ):
        batch_dataframe = record_batch.to_pandas()

        codes = encode_sequences_to_codes(
            sequences=batch_dataframe["sequence"].tolist(),
            property_table=property_table,
            expected_length=SEQUENCE_LENGTH,
        )

        shard_path = output_dir / f"{split_name}_physicochemical_{shard_index:05d}.npz"

        np.savez_compressed(
            shard_path,
            codes=codes,
            labels=batch_dataframe["label"].to_numpy(dtype=np.int8),
            chromosomes=batch_dataframe["chrom"].astype(str).to_numpy(dtype="U32"),
            positions=batch_dataframe["canonical_position"].to_numpy(dtype=np.int64),
        )

        shard_size = len(batch_dataframe)
        total_rows += shard_size

        shard_records.append(
            {
                "split": split_name,
                "shard_index": shard_index,
                "file": str(shard_path),
                "row_count": shard_size,
                "codes_shape": list(codes.shape),
                "dtype": str(codes.dtype),
            }
        )

        print(
            f"{split_name} | shard {shard_index:05d} | rows: {shard_size:,} | "
            f"total: {total_rows:,}"
        )

        shard_index += 1

    manifest_path = output_dir / "manifest.json"

    with manifest_path.open("w", encoding="utf-8") as file:
        json.dump(
            {
                "split": split_name,
                "source_parquet": str(parquet_path),
                "shard_size": PHYSICOCHEMICAL_SHARD_SIZE,
                "total_rows": total_rows,
                "shard_count": shard_index,
                "codes_shape_per_sample": [SEQUENCE_LENGTH - 1],
                "dtype": "uint8",
                "unknown_dinucleotide_code": UNKNOWN_DINUCLEOTIDE_CODE,
                "dinucleotide_codes": dinucleotide_codes(property_table),
                "note": "Use expand_codes_to_matrix() to reconstruct the [12, 500] float32 matrix at load time.",
                "shards": shard_records,
            },
            file,
            indent=2,
        )

    return {
        "split": split_name,
        "total_rows": total_rows,
        "shard_count": shard_index,
        "manifest": str(manifest_path),
    }


def main() -> None:
    PHYSICOCHEMICAL_FEATURES_DIR.mkdir(parents=True, exist_ok=True)

    property_table = load_physicochemical_properties_di(PHYSICOCHEMICAL_PROPERTY_FILE)
    print(f"Loaded physicochemical properties for {len(property_table)} dinucleotides.")

    summaries = []

    for split_name in SPLIT_NAMES:
        parquet_path = DATASET_DIR / f"{split_name}.parquet"

        if not parquet_path.exists():
            raise FileNotFoundError(
                f"{parquet_path} does not exist. Run preprocessing/preprocess_hepg2.py first."
            )

        print(f"\nProcessing split: {split_name}")
        summaries.append(process_split(split_name, parquet_path, property_table))

    summary_path = PHYSICOCHEMICAL_FEATURES_DIR / "extraction_summary.json"

    with summary_path.open("w", encoding="utf-8") as file:
        json.dump(summaries, file, indent=2)

    print("\nPhysicochemical feature extraction completed.")
    print(f"Summary: {summary_path}")

    for summary in summaries:
        print(f"  {summary['split']}: {summary['total_rows']:,} rows in {summary['shard_count']} shards")


if __name__ == "__main__":
    main()
